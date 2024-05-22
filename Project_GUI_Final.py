import PySimpleGUI as sg
import datetime
import openpyxl
import os
from sklearn.preprocessing import StandardScaler
from pandas.core.common import flatten
from variable_file import values_repository, sex_list, chestPainTypeList, fastinBS_list, restingECG_list, exerciseAngina_list, st_Slope_list, classifier_catboost, classifier_output_list, all_classifiers_op

sg.theme('DarkBrown2')

def create_or_update_excel(filename, data):
    if not os.path.exists(filename):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        for row in data:
            sheet.append(row)
        workbook.save(filename)
        print(f"Excel file '{filename}' created and values added.")
    else:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active if "Data" in workbook.sheetnames else workbook.create_sheet("Data")
        for row in data:
            sheet.append(row)
        workbook.save(filename)
        print(f"Values added to existing Excel file '{filename}'.")
        
def predict_and_generate_output(final_input_list):
    # output_data = classifier_catboost.predict([final_input_list])   #all_classifiers_op(classifier_output_list, final_input_list)
    output_data = all_classifiers_op(classifier_output_list, final_input_list)
    print(output_data)
    if output_data == 1:
        return 'The person has a high risk to have a heart disease.', 1 
    else:
        return 'The person does not has a high risk to have a heart disease.', 0

def input_formatter(patient_details_list, values_repository):
    parameter_dict_details = [int(values_repository[element]) if element in values_repository.keys() else int(element) for element in patient_details_list[2:]]
    # print(f'parameter details: {parameter_dict_details}\n')
    ss = StandardScaler()
    final_input_list = list()
    list_for_CSV = patient_details_list[0:2]
    list_for_CSV.extend(parameter_dict_details)
    # print(list_for_CSV)
    for i in range(len(list_for_CSV[2:])):
        if i in [0, 3, 4, 7]:
            final_input_list.append(ss.fit_transform([[parameter_dict_details[i]]]))
        else:
            final_input_list.append(parameter_dict_details[i])

    final_input_list = list(flatten(final_input_list))
    print("final_input_list")
    print(final_input_list)
    return final_input_list, list_for_CSV

def add_all_details(welcome_values,  details_values):
    welcome_values.update(details_values)
    # print(welcome_values)
    # parameter_list_details = [str(time_stamp_val)]
    parameter_list_details = list()
    parameter_list_details1 = [element for element in welcome_values.values()]
    parameter_list_details.extend(parameter_list_details1)
    # print(parameter_dict_details)
    return parameter_list_details

def goodbyeWindow():
    sg.popup("", "              GOODBYE!              \x00", "")
    
def NamePopUp():
    sg.popup("", "              PLEASE ENTER YOUR NAME!              \x00", "")
    
def PhonePopUp():
    sg.popup("", "              PLEASE ENTER YOUR PHONE NUMBER!                    \x00", "")

def BothPopUp():
    sg.popup("", "            PLEASE ENTER YOUR NAME AND PHONE NUMBER!                \x00", "")
    
def ValidPhonePopUp():
    sg.popup("", "            PLEASE ENTER A VALID PHOME NUMBER!              \x00", "")

def ValidMaxHRPopUp():
    sg.popup("", "            PLEASE ENTER A VALID MAX HEART RATE(60 to 202)              \x00", "")
    
def ValidRestingBP():
    sg.popup("", "            PLEASE ENTER A RESTING BLOOD PRESSURE              \x00", "")
    
def ValidCholesterol():
    sg.popup("", "            PLEASE ENTER A CHOLESTEROL LEVEL              \x00", "")
    
def ValidAge():
    sg.popup("", "            PLEASE ENTER YOUR AGE              \x00", "")

def EmptyHeart():
    sg.popup("", "            PLEASE ENTER YOUR HEART RATE              \x00", "")
   
def Prediction_Window(welcome_values, details_values, predicted_value, predicted_message):
    timestamp = datetime.datetime.now()
    predWin = sg.Window("PREDICTION WINDOW", size=(1200,800), margins=(0,50),
                           layout = [[sg.Text("                                          RESULT",font=("Times New Roman",32), justification='center')],
                                    [sg.Text(f"Name                 : {welcome_values['Name']}",font = ("Times New Roman",16), justification='left')],
                                    [sg.Text(f"Phone Number         : {welcome_values['Phone']}",font = ("Times New Roman",16), justification='left')],
                                    [sg.Text(f"Age                  : {details_values['Age']}", font = ("Times New Roman",16), justification='left')],
                                    [sg.Text(f"Sex                  : {details_values['Sex']}", font = ("Times New Roman",16), justification='left')],
                                    [sg.Text(f"Chest Pain Type      : {details_values['ChestPainType']}", font = ("Times New Roman",16), justification='left')],
                                    [sg.Text(f"Resting-BP           : {details_values['RestingBP']}", font = ("Times New Roman",16), justification='left')],
                                    [sg.Text(f"Cholesterol level    : {details_values['Cholesterol']}", font = ("Times New Roman",16), justification='left')],
                                    [sg.Text(f"Fasting Blood Pressure : {details_values['FastingBS']}", font = ("Times New Roman",16), justification='left')],
                                    [sg.Text(f"Resting-ECG          : {details_values['Resting_ECG']}", font = ("Times New Roman",16), justification='left')],
                                    [sg.Text(f"Mx Heart Rate        : {details_values['maxHR']}", font = ("Times New Roman",16), justification='left')],
                                    [sg.Text(f"Exercise-Angina      : {details_values['Exercise_Angina']}", font = ("Times New Roman",16), justification='left')],
                                    [sg.Text(f"ST-Slope : {details_values['ST_slope']}", font = ("Times New Roman",16), justification='left')],
                                    [sg.HorizontalSeparator()], 
                                    [sg.Text(predicted_message, font = ("Times New Roman", 19), justification='center')], 
                                    [sg.Button('EXIT',size=(13,1))]])
    while True:
        event, predvalues = predWin.read()
        if event == sg.WIN_CLOSED or event == 'CANCEL' or event == 'EXIT':
            break
        
    predWin.close()
    return timestamp

def DetailsWindow(welcome_values):
    detailsWin = sg.Window("HEART DISEASE PEREDICTION", size=(1200,900), margins=(10,100),
                           layout = [[sg.Text("              PLEASE FILL YOUR DETAILS BELOW",font=("Times New Roman",32), justification='center')],
                                     [sg.HorizontalSeparator()],
                                     [sg.Text(" AGE                                          :",font=("Times New Roman",16),justification='left',pad=(0,10)),sg.Input(key = 'Age', size=(26,10))],
                                     [sg.Text(" SEX                                           :",font=("Times New Roman",16),justification='left',pad=(0,10)),sg.DropDown(sex_list, size = (26, 10), default_value = sex_list[0], key = 'Sex')],
                                     [sg.Text("CHEST PAIN TYPE                :",font=("Times New Roman",16),justification='left'),sg.DropDown(chestPainTypeList, size = (25, 10), default_value = chestPainTypeList[0], key = 'ChestPainType')],
                                     [sg.Text("RESTING BP                           :",font=("Times New Roman",16),justification='left'),sg.Input(key = 'RestingBP', size=(26,10))],
                                     [sg.Text("CHOLESTEROL                     :",font=("Times New Roman",16),justification='left'),sg.Input(key = 'Cholesterol', size=(26,10))],
                                     [sg.Text(" FASTING BLOOD SUGAR    : ",font=("Times New Roman",16),justification='left',pad=(0,10)),sg.DropDown(fastinBS_list, size = (25, 10), default_value = fastinBS_list[0], key = 'FastingBS')],
                                     [sg.Text(" RESTING ECG                        : ",font=("Times New Roman",16),justification='left',pad=(0,10)),sg.DropDown(restingECG_list, size = (25, 10), default_value = restingECG_list[0], key = 'Resting_ECG')],
                                     [sg.Text(" MAXIMUM HEART RATE    : ",font=("Times New Roman",16),justification='left',pad=(0,10)),sg.Input(key = 'maxHR', size=(26,10))],
                                     [sg.Text(" EXERCISE-ANGINA              : ",font=("Times New Roman",16),justification='left',pad=(0,10)),sg.DropDown(exerciseAngina_list, size = (25, 10), default_value = exerciseAngina_list[0], key = 'Exercise_Angina')],
                                     [sg.Text(" ST SLOPE                                : ",font=("Times New Roman",16),justification='left',pad=(0,10)),sg.DropDown(st_Slope_list, size = (25, 10), default_value = st_Slope_list[0], key = 'ST_slope')],
                                     [sg.Button('SUBMIT',size=(13,1)),sg.Button('CANCEL',size=(13,1))]], finalize=True)
    flag = False
    while True:
        event, details_values = detailsWin.read()
        if event == sg.WIN_CLOSED or event == 'CANCEL':
            break
        elif event == 'SUBMIT':
            if details_values['Age'] == '':
                ValidAge()
            elif details_values['RestingBP'] == '':
                ValidRestingBP()
            elif details_values['Cholesterol'] == '':
                ValidCholesterol()
            elif details_values['maxHR'] == '':
                EmptyHeart()
            elif int(details_values['maxHR'])<60 or int(details_values['maxHR'])> 202 :
                ValidMaxHRPopUp()
            else:
                sg.popup('VALUES ENTERED SUCCESSFULLY')
                flag = True
                break
            
    detailsWin.close()
    if flag:
        list_for_input_formatter = add_all_details(welcome_values, details_values)
        print(list_for_input_formatter)
        final_input_list, list_for_CSV = input_formatter(list_for_input_formatter, values_repository)
        print(final_input_list)
        predicted_message, predicted_value = predict_and_generate_output(final_input_list)
        time_stamp = Prediction_Window(welcome_values, details_values, predicted_value, predicted_message)
    else:
        time_stamp = None
    return details_values, time_stamp, list_for_CSV

def welcomeWindow():
    welcomewin = sg.Window("HEART DISEASE PEREDICTION", size=(1200,400), margins=(0,50),
                           layout = [[sg.Text("WELCOME",font=("Times New Roman",32), justification='center')],
                                    [sg.Text("Please enter your initial details: ",font=("Times New Roman", 20,"bold"))],
                                    [sg.Text("NAME                   : ",font=("Times New Roman",16),justification='left',pad=(0,10)),sg.Input(key = 'Name', size=(70,10))],
                                    [sg.Text("MOBILE NUMBER : ",font=("Times New Roman",16),justification='left'),sg.Input(key = 'Phone', size=(70,10))],
                                    [sg.Button('SUBMIT',size=(13,1)),sg.Button('CANCEL',size=(13,1))]], element_justification='c')
    while True:
        event, welcome_values = welcomewin.read()
        if event == sg.WIN_CLOSED or event == 'CANCEL':
            break
        if event == 'SUBMIT':
            if welcome_values['Name']== "" and welcome_values['Phone']== "":
                BothPopUp()
            elif welcome_values['Name']== "":
                NamePopUp()
            elif welcome_values['Phone']== "":
                PhonePopUp()
            elif len(welcome_values['Phone'])!= 10:
                ValidPhonePopUp()
            else:
                details_values, time_stamp, list_for_CSV= DetailsWindow(welcome_values)
                break
        
    welcomewin.close()
    goodbyeWindow()
    return welcome_values, details_values, time_stamp, list_for_CSV
    
welcome_values,  details_values, time_stamp_val, list_for_CSV = welcomeWindow()

print(details_values)
list_for_CSV.insert(0, str(time_stamp_val))
create_or_update_excel("UserData.xlsx", [list_for_CSV])